from __future__ import annotations

import io
from pathlib import Path


def _html_to_pdf_bytes(html: str) -> bytes:
    import fitz

    story = fitz.Story(html)
    mediabox = fitz.paper_rect("a4")
    where = mediabox + (50, 50, -50, -50)

    buf = io.BytesIO()
    writer = fitz.DocumentWriter(buf)
    more = True
    while more:
        dev = writer.begin_page(mediabox)
        more, _ = story.place(where)
        story.draw(dev)
        writer.end_page()
    writer.close()
    return buf.getvalue()


def _xlsx_to_html(content: bytes) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    parts = ["<html><body>"]
    for name in wb.sheetnames:
        ws = wb[name]
        parts.append(f"<h2>{name}</h2><table border='1' cellpadding='4'>")
        for row in ws.iter_rows(values_only=True):
            parts.append("<tr>" + "".join(f"<td>{v or ''}</td>" for v in row) + "</tr>")
        parts.append("</table>")
    parts.append("</body></html>")
    return "\n".join(parts)


_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
_TEXT_EXTS = {".txt", ".md", ".csv", ".html", ".htm", ".json", ".xml", ".yaml", ".yml"}


def normalize(file_path: str | Path) -> tuple:
    """Normalize any supported file format to a fitz.Document.

    DOCX and PPTX use MuPDF's native converter (preserves formatting).
    XLSX uses openpyxl → HTML → PDF.
    Images are opened directly as single-page fitz documents.
    Text/markup formats are wrapped in a minimal HTML → PDF.

    Returns (fitz_doc, file_size_bytes). Caller must call doc.close().
    """
    import fitz

    path = Path(file_path)
    ext = path.suffix.lower()
    content = path.read_bytes()
    file_size = len(content)

    if ext == ".pdf":
        doc = fitz.open(stream=content, filetype="pdf")

    elif ext in {".docx", ".pptx"}:
        # MuPDF's native DOCX/PPTX reader preserves bold, italic, tables, images
        doc = fitz.open(path)
        try:
            pdf_bytes = doc.convert_to_pdf()
            doc.close()
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        except Exception:
            doc.close()
            raise

    elif ext == ".xlsx":
        html = _xlsx_to_html(content)
        pdf_bytes = _html_to_pdf_bytes(html)
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")

    elif ext in _IMAGE_EXTS:
        doc = fitz.open(stream=content, filetype=ext[1:])

    elif ext in _TEXT_EXTS:
        text = content.decode("utf-8", errors="replace")
        html = f"<html><body><pre>{text}</pre></body></html>"
        pdf_bytes = _html_to_pdf_bytes(html)
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")

    else:
        text = content.decode("utf-8", errors="replace")
        html = f"<html><body><pre>{text}</pre></body></html>"
        pdf_bytes = _html_to_pdf_bytes(html)
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")

    return doc, file_size
